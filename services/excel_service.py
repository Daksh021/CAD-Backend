"""
services/excel_service.py
Builds a formatted Excel workbook from balloon data.
Uses openpyxl with professional styling (column widths, colours, borders, freeze panes).
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ── Style constants ───────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1B3A5C")   # dark navy
ALT_FILL      = PatternFill("solid", fgColor="EFF4FB")   # light blue-grey
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT     = Font(name="Calibri", size=10)
CENTRE        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN          = Side(style="thin",   color="B0BEC5")
THICK         = Side(style="medium", color="1B3A5C")

_TYPE_COLOUR = {
    "dimension":    "1565C0",   # blue
    "tolerance":    "6A1B9A",   # purple
    "gdt":          "00695C",   # teal
    "surface_finish":"E65100",  # orange
    "note":         "37474F",   # dark grey
    "specification":"558B2F",   # green
}


def _cell_border(top=THIN, right=THIN, bottom=THIN, left=THIN):
    return Border(top=top, right=right, bottom=bottom, left=left)


# ── Public API ────────────────────────────────────────────────────────────────

def build_excel_workbook(drawing, balloons, include_remarks: bool = True) -> Workbook:
    """
    Create and return an openpyxl Workbook containing:
    - Sheet 1: Balloon inspection table
    - Sheet 2: Summary / metadata
    """
    wb = Workbook()

    _build_balloon_sheet(wb, drawing, balloons, include_remarks)
    _build_summary_sheet(wb, drawing, balloons)

    return wb


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_balloon_sheet(wb: Workbook, drawing, balloons, include_remarks: bool):
    ws = wb.active
    ws.title = "Balloon Inspection"

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value    = f"Balloon Inspection Report — {drawing.original_name}"
    title_cell.font     = Font(name="Calibri", bold=True, size=14, color="1B3A5C")
    title_cell.alignment = CENTRE
    ws.row_dimensions[1].height = 28

    # ── Column headers ───────────────────────────────────────────────────────
    headers = [
        ("Balloon No.", 12),
        ("Page No.",    10),
        ("Type",        16),
        ("Drawing Reference / Text", 35),
        ("Description", 40),
        ("X %",         8),
        ("Y %",         8),
    ]
    if include_remarks:
        headers.append(("Remarks", 30))

    for col_idx, (label, width) in enumerate(headers, start=1):
        cell            = ws.cell(row=2, column=col_idx, value=label)
        cell.fill       = HEADER_FILL
        cell.font       = HEADER_FONT
        cell.alignment  = CENTRE
        cell.border     = _cell_border(top=THICK, bottom=THICK)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, b in enumerate(balloons, start=3):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL

        type_colour = _TYPE_COLOUR.get(b.balloon_type.value if hasattr(b.balloon_type, "value") else str(b.balloon_type), "37474F")

        row_data = [
            b.balloon_number,
            b.page_number,
            (b.balloon_type.value if hasattr(b.balloon_type, "value") else str(b.balloon_type)).upper(),
            b.extracted_text or "",
            b.description   or "",
            f"{b.x_pct:.1f}",
            f"{b.y_pct:.1f}",
        ]
        if include_remarks:
            row_data.append(b.remarks or "")

        for col_idx, value in enumerate(row_data, start=1):
            cell            = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font       = Font(name="Calibri", size=10,
                                   color=type_colour if col_idx == 3 else "212121",
                                   bold=(col_idx == 1))
            cell.fill       = fill
            cell.border     = _cell_border()
            cell.alignment  = CENTRE if col_idx in (1, 2, 6, 7) else LEFT

        ws.row_dimensions[row_idx].height = 18

    # ── Autofilter on header row ─────────────────────────────────────────────
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A2:{last_col}2"


def _build_summary_sheet(wb: Workbook, drawing, balloons):
    ws = wb.create_sheet("Summary")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28

    summary_data = [
        ("Drawing",        drawing.original_name),
        ("Total Pages",    drawing.page_count),
        ("Total Balloons", len(balloons)),
        ("Auto-detected",  sum(1 for b in balloons if b.is_auto)),
        ("Manual",         sum(1 for b in balloons if not b.is_auto)),
    ]

    # Add per-type breakdown
    from collections import Counter
    type_counts = Counter(
        (b.balloon_type.value if hasattr(b.balloon_type, "value") else str(b.balloon_type))
        for b in balloons
    )
    for btype, count in type_counts.items():
        summary_data.append((f"  — {btype.title()}", count))

    ws.merge_cells("A1:B1")
    ws["A1"].value     = "Summary"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=13, color="1B3A5C")
    ws["A1"].alignment = CENTRE
    ws["A1"].fill      = PatternFill("solid", fgColor="D6E4F0")

    for i, (label, value) in enumerate(summary_data, start=2):
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=value)
        lc.font = Font(name="Calibri", size=10, bold=True)
        vc.font = Font(name="Calibri", size=10)
        lc.fill = vc.fill = (ALT_FILL if i % 2 == 0 else WHITE_FILL)
        lc.alignment = LEFT
        vc.alignment = CENTRE
        lc.border = vc.border = _cell_border()